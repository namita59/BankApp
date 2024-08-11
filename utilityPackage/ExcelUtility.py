import openpyxl

def getRowCount(file,sheetname):
    book =openpyxl.load_workbook(file)
    sheet =book[sheetname]
    return sheet.max_row

def readData(file,sheetname,rowno,colmno):
    book = openpyxl.load_workbook(file)
    sheet =book[sheetname]
    return sheet.cell(row =rowno,column=colmno).value

def writeData(file,sheetname,rowmo,columno,data):
    book =openpyxl.load_workbook(file)
    sheet =book[sheetname]
    sheet.cell(row=rowmo,column=columno).value=data
    book.save(file)
